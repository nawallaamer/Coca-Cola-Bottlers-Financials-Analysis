import pandas as pd
import openpyxl

#Print Statements and Input 
print("Welcome to the Analysis by Nawall. ")
# ticker= input("Enter the Required Company: ")
# ticker = str(ticker)
# date = input("Enter the Required Year: ")
# date=str(date)
tickers = ['AC.MX','AEFES.E','CCOLA.E','KO','KOF','CCHGY','BUD','PEP','CCEP','AKOA']
dates = ['2022','2021','2020','2019', '2018']
for ticker in tickers:
    for date in dates:
        #Reading the Data
        data_bs = pd.read_excel(f'Financial Statements ({ticker}).xlsx',sheet_name='Balance Sheet')
        data_is= pd.read_excel(f'Financial Statements ({ticker}).xlsx',sheet_name='Income Statement')
        data_cf= pd.read_excel(f'Financial Statements ({ticker}).xlsx',sheet_name='Cash Flow')

        # date_column1 = data_bs.columns[data_bs.columns.str.endswith(date)][0]
        # date_column2 = data_is.columns[data_is.columns.str.endswith(date)][0]
        # date_column3 = data_cf.columns[data_cf.columns.str.endswith(date)][0]

        #Setting the Values in the Breakdown Column as Indexes 
        data_bs.set_index(data_bs.columns[0], inplace=True)
        data_is.set_index(data_is.columns[0], inplace=True)
        data_cf.set_index(data_cf.columns[0], inplace=True)

        #Variables and Converting into INT

        profit= data_is.loc['Net Income',date]
        revenue= data_is.loc['Sales/Revenue',date]
        ebit= data_is.loc['EBIT',date]
        ebit=ebit[0]
        gross_profit= data_is.loc['Gross Income',date]
        operating_cash_flow= data_cf.loc['Net Operating Cash Flow',date]
        sales_growth= data_is.loc['Sales Growth',date]
        gross_income_growth= data_is.loc['Gross Income Growth',date]

        if type(profit) == str:
            profit=profit.replace(',','')
            profit= float(profit)
        else:
            profit= float(profit)
        if type(revenue) == str:
            revenue=revenue.replace(',','')
            revenue= float(revenue)
        else:
            revenue= float(revenue)
        if type(ebit) == str:
            ebit=ebit.replace(',','')
            ebit= float(ebit)
        else:
            ebit= float(ebit)
        if type(gross_profit) == str:
            gross_profit=gross_profit.replace(',','')
            gross_profit= float(gross_profit)
        else:
            gross_profit= float(gross_profit)
        if type(operating_cash_flow) == str:
            operating_cash_flow=operating_cash_flow.replace(',','')
            operating_cash_flow= float(operating_cash_flow)
        else:
            operating_cash_flow= float(operating_cash_flow)

        if ticker=='AC.MX' or ticker=='CCOLA.E' or ticker=='AEFES.E' or ticker== 'AKOA' or ticker== 'BUD' or ticker== 'CCHGY' or ticker== 'KOF':
            current_assets= data_bs.loc['Total Current Assets',date]
            inventory= data_bs.loc['Inventories',date]
            current_liabilities= data_bs.loc['Total Current Liabilities',date]
            total_liabilities= data_bs.loc['Total Liabilities',date]
            shareholders_equity= data_bs.loc["Total Shareholders' Equity",date]
            outstanding_share= data_bs.loc['Common Stock Par/Carry Value',date]
            earnings= data_bs.loc['Retained Earnings',date]
            current_debt= data_bs.loc['ST Debt & Current Portion LT Debt',date]
            long_debt= data_bs.loc['Long-Term Debt',date]
            total_assets=data_bs.loc['Total Assets',date]
        else:
            current_assets= data_bs.loc['Total Current Assets',date]
            current_assets=current_assets.replace(',','')
            current_assets= float(current_assets)
            inventory= data_bs.loc['Inventories',date]
            inventory=inventory.replace(',','')
            inventory= float(inventory)
            current_liabilities= data_bs.loc['Total Current Liabilities',date]
            current_liabilities=current_liabilities.replace(',','')
            current_liabilities= float(current_liabilities)
            total_liabilities= data_bs.loc['Total Liabilities',date]
            total_liabilities=total_liabilities.replace(',','')
            total_liabilities= float(total_liabilities)
            shareholders_equity= data_bs.loc["Total Shareholders' Equity",date]
            shareholders_equity=shareholders_equity.replace(',','')
            shareholders_equity= float(shareholders_equity)
            outstanding_share= data_bs.loc['Common Stock Par/Carry Value',date]
            outstanding_share=outstanding_share.replace(',','')
            outstanding_share= float(outstanding_share)
            earnings= data_bs.loc['Retained Earnings',date]
            earnings=earnings.replace(',','')
            earnings= float(earnings)
            current_debt= data_bs.loc['ST Debt & Current Portion LT Debt',date]
            current_debt=current_debt.replace(',','')
            current_debt= float(current_debt)
            long_debt= data_bs.loc['Long-Term Debt',date]
            long_debt=long_debt.replace(',','')
            long_debt= float(long_debt)
            total_assets=data_bs.loc['Total Assets',date]
            total_assets=total_assets.replace(',','')
            total_assets= float(total_assets)

        roa=data_bs.loc['Return On Average Assets',date]

        #Calculating Ratios
        Return_on_Equity= profit/shareholders_equity
        Quick_Ratio= (current_assets-inventory)/current_liabilities
        Debt_to_Equity_Ratio= total_liabilities/shareholders_equity
        Working_Capital_Ratio= current_assets/current_liabilities
        Earning_Per_Share= profit/outstanding_share
        Net_Profit_Margin= profit/revenue
        EBIT_Margin= (ebit/revenue)*100
        Gross_Profit_Margin= gross_profit/revenue
        Cash_Flow_Margin= operating_cash_flow/revenue
        Leverage_Ratio= (current_debt+long_debt)/total_assets

        #Printing the Ratios 
        print('Return on Equity: ')
        print(round(Return_on_Equity, 3))
        print('Quick Ratio: ')
        print(round(Quick_Ratio, 3))
        print('Debt to Equity Ratio: ')
        print(round(Debt_to_Equity_Ratio, 3))
        print('Working Capital Ratio: ')
        print(round(Working_Capital_Ratio, 3))
        print('Earning Per Share: ')
        print(round(Earning_Per_Share, 3))
        print('Net Profit Margin: ')
        print(round(Net_Profit_Margin, 3))
        print('EBIT Margin: ')
        print(round(EBIT_Margin, 3))
        print('Gross Profit Margin: ')
        print(round(Gross_Profit_Margin, 3))
        print('Cash Flow Margin: ')
        print(round(Cash_Flow_Margin, 3))
        print('Leverage Ratio: ')
        print(round(Leverage_Ratio, 3))
        print('Sales Growth: ')
        print(sales_growth)
        print('Gross Income Growth: ')
        print(gross_income_growth)
        print('Return on Assets: ')
        print(roa)


        #Saving in Excel
        ratios=['Return on Equity', 'Quick Ratio', 'Debt to Equity Ratio', 'Working Capital Ratio', 'Earning Per Share', 'Net Profit Margin', 'EBIT Margin', 'Gross Profit Margin', 'Cash Flow Margin', 'Leverage Ratio', 'Sales Growth', 'Gross Income Growth', 'Return on Assets']
        values=[round(Return_on_Equity,3), round(Quick_Ratio,3), round(Debt_to_Equity_Ratio,3), round(Working_Capital_Ratio,3), round(Earning_Per_Share,3), round(Net_Profit_Margin,3), round(EBIT_Margin,3), round(Gross_Profit_Margin,3), round(Cash_Flow_Margin,3), round(Leverage_Ratio,3), sales_growth, gross_income_growth, roa]

        workbook = openpyxl.load_workbook(f'Financial Statements ({ticker}).xlsx')
        if 'Ratio' in workbook.sheetnames:
            analysis_sheet = workbook['Ratio']
        else:
            analysis_sheet = workbook.create_sheet(title='Ratio')

        analysis_sheet['A1'] = 'Analysis Ratios'
        analysis_sheet['B1'] = '2022'
        analysis_sheet['C1'] = '2021'
        analysis_sheet['D1'] = '2020'
        analysis_sheet['E1'] = '2019'
        analysis_sheet['F1'] = '2018'

        for r in range(len(ratios)):
            analysis_sheet.cell(row=(r+2),column=1, value=ratios[r])
        if date=='2022':
            for k in range(len(values)):
                analysis_sheet.cell(row=(k+2), column=2, value=values[k])
        if date=='2021':
            for k in range(len(values)):
                analysis_sheet.cell(row=(k+2), column=3, value=values[k])
        if date=='2020':
            for k in range(len(values)):
                analysis_sheet.cell(row=(k+2), column=4, value=values[k])
        if date=='2019':
            for k in range(len(values)):
                analysis_sheet.cell(row=(k+2), column=5, value=values[k])
        if date=='2018':
            for k in range(len(values)):
                analysis_sheet.cell(row=(k+2), column=6, value=values[k])

        for column in analysis_sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            analysis_sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        workbook.save(f'Financial Statements ({ticker}).xlsx')